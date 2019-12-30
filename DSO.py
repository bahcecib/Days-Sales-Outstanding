# Written by Berker Bahceci, July&August 2019
# Sanofi CI2C Turkey
# DSO-Health report automation

import pandas as pd
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf
from calendar import monthrange
import tkinter as tk
from tkinter import filedialog
import time
start = time.time()






###############################################################################
################### PART 1: ANALYSIS ##########################################
###############################################################################


# Get all the necessary file paths
root=tk.Tk()
root.lift()
root.attributes("-topmost", True)
aging_file_path = filedialog.askopenfilename(title='Select the SAP aging data')
customer_dict_file_path = filedialog.askopenfilename(title='Select customer dictionnary')
report_path = filedialog.asksaveasfilename(title='Save the report to', filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*") ))
sales_file_path = filedialog.askopenfilename(title='Select the SAP sales data')
sales_db_path = filedialog.askopenfilename(title='Select the sales database')
root.withdraw()




# Raw aging data
raw_age = pd.read_excel(aging_file_path)
raw_header = list(raw_age.columns)


# Credit limits
kredi_limiti = raw_age[[raw_header[8],raw_header[9],raw_header[19]]].copy()
kredi_limiti.drop_duplicates(inplace=True)
kredi_limiti = kredi_limiti.reset_index(drop=True)
kredi_limiti = kredi_limiti.dropna(axis=0, how='any')
kredi_limiti.rename(columns={raw_header[9]:'Müsteri'}, inplace=True)




# Grouped LOG table
# Each customer has three rows of LOG: ÖDK={B,M or X}
log = raw_age[[raw_header[8], raw_header[15], raw_header[22]]].copy()
log.fillna('X', inplace=True)
log = log.groupby([raw_header[8], raw_header[22]]).sum().reset_index() 




# Initialize the summation /customer group of LOG and OAR
log_sum = pd.DataFrame(columns=['Müsteri Kodu', 'LOG'])
log_sum['Müsteri Kodu'] = kredi_limiti[raw_header[8]].copy()
log_sum.fillna(0, inplace=True)

risk_sum = pd.DataFrame(columns=['Müsteri Kodu', 'Total risk'])
risk_sum['Müsteri Kodu'] = kredi_limiti[raw_header[8]].copy()
risk_sum.fillna(0, inplace=True)




# Created LOG and OAR in seperate dataframes
# Get the data from those dataframes to the new, summed frames
# Reduce 3 rows of LOG to a single one
for i in range(len(log)):
    
    if log.iloc[i,1] == 'M':
        for j in range(len(log_sum)):
            if log.iloc[i,0] == log_sum.iloc[j,0]:
                log_sum.iloc[j,1] = log.iloc[i,2]
    else:
        for j in range(len(risk_sum)):
            if log.iloc[i,0] == risk_sum.iloc[j,0]:
                risk_sum.iloc[j,1] = risk_sum.iloc[j,1]+log.iloc[i,2]




# Inıtialize dataframe to write the due date data         
vade_header = [raw_header[8], raw_header[22], raw_header[17], raw_header[18], raw_header[23], raw_header[24], raw_header[25], raw_header[26], raw_header[27], raw_header[28]]       
due_raw = raw_age[vade_header].copy()




# Drop drows with ÖDK='M'
# ÖDK='M' was only used in LOG calculation. They are used there. 
# They shouldn't be taken into account for OAR and due date calculations.
indexNames = due_raw[due_raw[raw_header[22]] == 'M' ].index
due_raw.drop(indexNames, inplace=True)                
due = pd.DataFrame(columns=['Müsteri Kodu1', 'Overdue','Not overdue','0-30','30-60','60-90','90-120','120-150','>150'])
due['Müsteri Kodu1'] = kredi_limiti[raw_header[8]].copy()
due.fillna(0, inplace=True)


for m in range(len(due)):
    for k in range(len(due_raw)):
        if due_raw.iloc[k,0]==due.iloc[m,0]:
            due.iloc[m,1] = due.iloc[m,1]+due_raw.iloc[k,2]
            due.iloc[m,2] = due.iloc[m,2]+due_raw.iloc[k,3]
            due.iloc[m,3] = due.iloc[m,3]+due_raw.iloc[k,4]
            due.iloc[m,4] = due.iloc[m,4]+due_raw.iloc[k,5]
            due.iloc[m,5] = due.iloc[m,5]+due_raw.iloc[k,6]
            due.iloc[m,6] = due.iloc[m,6]+due_raw.iloc[k,7]
            due.iloc[m,7] = due.iloc[m,7]+due_raw.iloc[k,8]
            due.iloc[m,8] = due.iloc[m,8]+due_raw.iloc[k,9]
            



# There are negative overdue values in the data
# Take the negative overdues and add them to Not overdue, 0-30 days, without changing the sign
for m in range(len(due)):
    if due.iloc[m,1]<0:
        x = due.iloc[m,1]
        due.iloc[m,1] = 0
        due.iloc[m,2] = due.iloc[m,2]+x
        due.iloc[m,3] = due.iloc[m,3]+x




# All the data is now in the table with a single row for each customer
# Create a final dataframe with the total analysis details (Customer, OAR, Due, LOG, Credit Limit)
final = pd.concat([kredi_limiti,log_sum['LOG'],risk_sum['Total risk'],due], axis=1)
final.drop(['Müsteri Kodu1'], axis=1, inplace=True)
final.rename(columns={raw_header[8]:'Müsteri Kodu'}, inplace=True)
final.loc['Total'] = final.sum(numeric_only=True, axis=0)




# Assign customers their customer groups so that group dataframes to be written into different sheets
# This part reaches to the customer dictionnary and assigns groups
final['Group']=""
customer_dictionnary = pd.read_excel(customer_dict_file_path)
names = list(customer_dictionnary.columns)
size = customer_dictionnary.shape # Get the shape of the dataframe as a variable. So if size changes, code still works
for i in range(size[0]):
    for j in range(size[1]):
        for k in range(len(final)-1):
            if final.iloc[k,0] == customer_dictionnary.iloc[i,j]:
                final.iloc[k,13] = names[j]
                
contract=final[final.Group == 'Contract']
contract.loc['Total']=contract.sum(numeric_only=True, axis=0)
contract.loc['Total','Müsteri']='Total'
contract.drop('Group', axis=1, inplace=True)

alliance=final[final.Group == 'alliance']
alliance.loc['Total']=alliance.sum(numeric_only=True, axis=0)
alliance.loc['Total','Müsteri']='Total'
alliance.drop('Group', axis=1, inplace=True)

asgroup=final[final.Group == 'as']
asgroup.loc['Total']=asgroup.sum(numeric_only=True, axis=0)
asgroup.loc['Total','Müsteri']='Total'
asgroup.drop('Group', axis=1, inplace=True)

bursa=final[final.Group == 'bursa']
bursa.drop('Group', axis=1, inplace=True)

dilek=final[final.Group == 'dilek']
dilek.drop('Group', axis=1, inplace=True)

gek=final[final.Group == 'guney']
gek.drop('Group', axis=1, inplace=True)

istkop=final[final.Group == 'ist.koop.']
istkop.drop('Group', axis=1, inplace=True)

nevzat=final[final.Group == 'nevzat']
nevzat.loc['Total']=nevzat.sum(numeric_only=True, axis=0)
nevzat.loc['Total','Müsteri']='Total'
nevzat.drop('Group', axis=1, inplace=True)

other=final[final.Group == 'other wholesalers']
other.loc['Total']=other.sum(numeric_only=True, axis=0)
other.loc['Total','Müsteri']='Total'
other.drop('Group', axis=1, inplace=True)

selcuk=final[final.Group == 'selcuk']
selcuk.loc['Total']=selcuk.sum(numeric_only=True, axis=0)
selcuk.loc['Total','Müsteri']='Total'
selcuk.drop('Group', axis=1, inplace=True)

"""tender=final[final['Müsteri Kodu'].isin(customer_dictionnary['tender'])==True]
tender.loc['Total']=tender.sum(numeric_only=True, axis=0)
tender.loc['Total','Müsteri']='Total'
tender.drop('Group', axis=1, inplace=True)"""

other_tender=final[final.Group == 'other tender']
other_tender.loc['Total']=other_tender.sum(numeric_only=True, axis=0)
other_tender.loc['Total','Müsteri']='Total'
other_tender.drop('Group', axis=1, inplace=True)

aksel=final[final.Group == 'aksel']
aksel.drop('Group', axis=1, inplace=True)

ozsel=final[final.Group == 'ozsel']
ozsel.loc['Total']=ozsel.sum(numeric_only=True, axis=0)
ozsel.loc['Total','Müsteri']='Total'
ozsel.drop('Group', axis=1, inplace=True)

final.drop('Group', axis=1, inplace=True)
customers = [contract, alliance, asgroup, bursa, dilek, gek, istkop, nevzat, other, selcuk, final, other_tender, aksel, ozsel]



# The report template gets written
# Each customer has a sheet with the LOG, Credit Limit, OAR and Due data
with pd.ExcelWriter('%s.xlsx' % report_path) as writer:
    contract.to_excel(writer, sheet_name='Contract', index=False, header=final.keys())
    alliance.to_excel(writer, sheet_name='Alliance', index=False, header=final.keys())
    asgroup.to_excel(writer, sheet_name='AS Group', index=False, header=final.keys())
    bursa.to_excel(writer, sheet_name='Bursa', index=False, header=final.keys())
    dilek.to_excel(writer, sheet_name='Dilek', index=False, header=final.keys())
    gek.to_excel(writer, sheet_name='Güney Ecza', index=False, header=final.keys())
    istkop.to_excel(writer, sheet_name='Ist.Koop.', index=False, header=final.keys())
    nevzat.to_excel(writer, sheet_name='Nevzat Group', index=False, header=final.keys())
    other.to_excel(writer, sheet_name='Other Wholesalers', index=False, header=final.keys())
    selcuk.to_excel(writer, sheet_name='Selçuk Group', index=False, header=final.keys())
    final.to_excel(writer, sheet_name='All customers', index=False, header=final.keys())
   # tender.to_excel(writer, sheet_name='Tender', index=False, header=final.keys())
    other_tender.to_excel(writer, sheet_name='Other tender', index=False, header=final.keys())
    aksel.to_excel(writer, sheet_name='Aksel', index=False, header=final.keys())
    ozsel.to_excel(writer, sheet_name='Özsel', index=False, header=final.keys())






###############################################################################
################### PART 2: SALES #############################################
###############################################################################
    

# Get the Excel file which has the monthly sales data
raw_sales = pd.read_excel(sales_file_path) #, sheet_name='SAP sales July.19') #Again, talk to Dilek and get rid of sheet name
raw_sales = raw_sales.drop(len(raw_sales)-1, axis=0)




# Remove unnecessary columns&calculate sum
raw_sales_header = list(raw_sales.columns)
sales_header = [raw_sales_header[8], raw_sales_header[13], raw_sales_header[14], raw_sales_header[2]]
sales = raw_sales[sales_header].copy()
sales.rename(columns={raw_sales_header[8]:'Tutar', raw_sales_header[14]:'Müsteri Adi', raw_sales_header[2]:'Tarih'}, inplace=True)
grouped_sales = sales.groupby(['Müsteri Adi','Hesap','Tarih']).sum().reset_index()
grouped_sales['OAR'] = ""




# Assign customers their customer groups so that group dataframes to be written into different sheets
# This part reaches to the customer dictionnary and assigns groups
group_names = list(customer_dictionnary.columns)
grouped_sales['Group'] = ""

size = customer_dictionnary.shape # Get the shape of the dataframe as a variable. So if size changes, code still works
for i in range(size[0]):
    for j in range(size[1]):
        for k in range(len(grouped_sales)):
           if grouped_sales.iloc[k,1] == customer_dictionnary.iloc[i,j]:
                grouped_sales.iloc[k,5] = group_names[j]




# Add a totals row for cumulative sales
# This total cumulative sales will be written to update the Sales,OAR&DSO database
grouped_sales.loc['Total'] = grouped_sales.sum(numeric_only=True, axis=0)
grouped_sales.loc['Total','OAR'] = final.loc['Total','Total risk']
grouped_sales.loc['Total','Müsteri Adi'] = 'Total'
grouped_sales.rename(columns={'Hesap':'Müsteri Kodu'}, inplace=True)
grouped_sales.loc['Total','Group'] = 'Total'




"""#Aksel and Ozsel are written as tender but we need them as seperate customers
#Duplicate their rows and write groups as Aksel and Ozsel
#This is done after total sales row is added because otherwise their data would be added twice
row_aksel=grouped_sales[grouped_sales['Müsteri Kodu']==10167155.0].copy().reset_index()
row_aksel.drop('index', axis=1, inplace=True)
row_aksel['Group']='aksel'          
row_ozsel1=grouped_sales[grouped_sales['Müsteri Kodu']==10008657.0].copy().reset_index()
row_ozsel1.drop('index', axis=1, inplace=True)
row_ozsel1['Group']='ozsel' 
row_ozsel2=grouped_sales[grouped_sales['Müsteri Kodu']==10009513.0].copy().reset_index()
if row_ozsel2.empty==False:
    row_ozsel2=grouped_sales[grouped_sales['Müsteri Kodu']==10009513.0].copy().reset_index()
    row_ozsel2.drop('index', axis=1, inplace=True)
    row_ozsel2['Group']='ozsel' 
    row_ozsel=pd.concat([row_ozsel1, row_ozsel2], ignore_index=True)
    del[[row_ozsel1, row_ozsel2]]
else:
    row_ozsel=row_ozsel1




#Concatenate Aksel and Ozsel to the grouped sales data, which is also handled and cleaned
#Finally, the small dataframe of each customer group with monthly sales and an empty OAR column
grouped_sales=pd.concat([grouped_sales, row_aksel, row_ozsel], ignore_index=True)"""
grouped_sales.drop('Müsteri Kodu', axis=1, inplace=True)
dso_sales_final = grouped_sales.groupby('Group').sum(numeric_only=True)
new_index = ['Contract', 'alliance', 'as', 'bursa', 'dilek', 'guney', 'ist.koop.', 'nevzat', 'other wholesalers', 'selcuk', 'Total', 'other tender', 'aksel', 'ozsel']
dso_sales_final = dso_sales_final.reindex(new_index)


# Fill the OAR column with the data from Part-1
dso_sales_final['OAR']=""
dso_sales_final.loc['Contract','OAR'] = contract.loc['Total','Total risk']
dso_sales_final.loc['alliance','OAR'] = alliance.loc['Total','Total risk']
dso_sales_final.loc['as','OAR'] = asgroup.loc['Total','Total risk']
dso_sales_final.loc['bursa','OAR'] = bursa.iloc[0,4]
dso_sales_final.loc['dilek','OAR'] = dilek.iloc[0,4]
dso_sales_final.loc['guney','OAR'] = gek.iloc[0,4]
dso_sales_final.loc['ist.koop.','OAR'] = istkop.iloc[0,4]
dso_sales_final.loc['nevzat','OAR'] = nevzat.loc['Total','Total risk']
dso_sales_final.loc['other wholesalers','OAR'] = other.loc['Total','Total risk']
dso_sales_final.loc['selcuk','OAR'] = selcuk.loc['Total','Total risk']
dso_sales_final.loc['Total','OAR'] = final.loc['Total','Total risk']
dso_sales_final.loc['other tender','OAR'] = other_tender.loc['Total','Total risk']
dso_sales_final.loc['aksel','OAR'] = aksel.iloc[0,4]
dso_sales_final.loc['ozsel','OAR'] = ozsel.loc['Total','Total risk']






##############################################################################
################### PART 3: UPDATE THE REPORT TEMPLATE AND DATABASES #########
##############################################################################


# Update the database file with this month's Sales&OAR
wb = openpyxl.load_workbook(sales_db_path)
ws = wb.active
max_columns = ws.max_column
max_rows = ws.max_row
while ws.cell(row=max_rows, column=2).value == None:
        ws.delete_rows(max_rows,1)
        max_rows = ws.max_row




for i in range(2, max_columns, 3):
   ws.cell(row = max_rows + 1, column = i).value = dso_sales_final.iloc[int(np.floor(i/3)),0]

for i in range(3, max_columns + 1, 3):
   ws.cell(row = max_rows + 1, column = i).value = dso_sales_final.iloc[int(i/3)-1,1]

ws.cell(row = max_rows + 1, column = 1).value=grouped_sales.iloc[0,1]
     
max_columns = ws.max_column
max_rows = ws.max_row
for i in range(1, max_rows+5):
    if ws.cell(row=i, column=2).value == None:
        ws.delete_rows(i,1)
wb.save(sales_db_path)




# Load the updated sales databse and calculate the DSO
dso_table = pd.read_excel(sales_db_path)  #If it gives an error here, use %s %
length_dso = len(dso_table)
for i in range(length_dso-12):
    dso_table.drop(i, axis=0, inplace=True)
    
dso_table = dso_table.reset_index()
dso_table.drop('index', axis=1, inplace=True)


# DSO calculator
for i in range (3, max_columns,3):
    c = 0
    days = 0
    dso = dso_table.iloc[11,i-1]
    for j in range(11,5,-1):
        if dso-dso_table.iloc[j,i-2]>0:
            dso = dso-dso_table.iloc[j,i-2]
            days = days+monthrange(dso_table.iloc[j,0].year, dso_table.iloc[j,0].month)[1]
            c = c+1
        else: 
            break
    
    dso_table.iloc[11,i] = (dso/dso_table.iloc[j,i-2])*30+days


# Write calculated DSO value back into sales database.    
wb = openpyxl.load_workbook(sales_db_path)
ws = wb.active
max_row = ws.max_row
for i in range(3,max_columns,3):
    ws.cell(row=max_row, column=i+1).value = dso_table.iloc[len(dso_table)-1, i]
wb.save(sales_db_path)





# Update the report template with the DSO value for this month which is calculated above
i=0;j=0
wb = openpyxl.load_workbook('%s.xlsx' %report_path)
sheets = wb.sheetnames
for i in range(len(sheets)):
    ws = wb[sheets[i]]
    max_rows = ws.max_row
    max_columns = ws.max_column
    for j in range(len(dso_table)):
        #Give context to the DSO table with this.
        ws.cell(row=1, column=max_columns+4).value = 'DSO Table'
        ws.cell(row=2, column=max_columns+4).value = 'Monthly sales'
        ws.cell(row=2, column=max_columns+5).value = 'Open account risk'
        ws.cell(row=2, column=max_columns+6).value = 'DSO'
        ws.cell(row=2, column=max_columns+7).value = 'Period'
        
        ws.cell(row=j+3, column=max_columns+4).value = dso_table.iloc[j,3*(i+1)-2]
        ws.cell(row=j+3, column=max_columns+5).value = dso_table.iloc[j,3*(i+1)-1]
        ws.cell(row=j+3, column=max_columns+6).value = dso_table.iloc[j,3*(i+1)]
        ws.cell(row=j+3, column=max_columns+7).value = dso_table.iloc[j,0]



wb.save('%s.xlsx' % report_path)





##############################################################################
################### PART 4: FINALIZE REPORT WITH CHARTS ######################
##############################################################################

# Fixed values to be used in each chart is calculated before the loop for efficiency
wb = openpyxl.load_workbook('%s.xlsx' % report_path)
sheets = wb.sheetnames
all_customer_dso=dso_table['Unnamed: 33']
period=dso_table['Unnamed: 0'].dt.to_period('M')
credit_lim_total = final.iloc[len(final)-1, 2]   # Check if iloc values are true
ar_total = final.iloc[len(final)-1, 4]           # Check if iloc values are true
pos = list(range(len(all_customer_dso)))
width = 0.25
min1 = min(all_customer_dso)
max1 = max(all_customer_dso)
pdf = matplotlib.backends.backend_pdf.PdfPages("Charts PDF output.pdf")


for i in range(0, len(sheets)):
    if sheets[i]=='All customers':
        continue
    
    ws = wb[sheets[i]]
    max_rows = ws.max_row
    max_columns = ws.max_column




    # Sales pie
    # Labels and values for the chart
    labels = [sheets[i], 'Rest of the sales']
    x = dso_sales_final.iloc[i,0]/dso_sales_final.loc['Total', 'Tutar']
    sizes = [x, 1-x]
    
    explode = (0.1, 0) 
    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, explode=explode, colors=['tan', 'lightblue'], autopct='%1.1f%%', startangle=15)
    ax1.axis('equal')
    plt.figure(figsize=(4,2))
    
    
    # Title and legend
    ax1.set_title("%s's Monthly Sales Volume" %sheets[i])
    ax1.legend(labels, loc='best', bbox_to_anchor=(0.3, 0.2,))
    fig1.savefig("sales_pie%s.png" % i, dpi = 72)
    pdf.savefig(fig1)


    

    # DSO bar
    fig2, ax2 = plt.subplots(figsize=(14,8))
    # Insert data into bars
    plt.bar(pos, all_customer_dso,width, label='All customers', color='lightblue')
    plt.bar([p + width for p in pos], dso_table.iloc[:,3*i+3], width, label='%s' % sheets[i], color='tan')


    # Labels
    ax2.set_ylabel('DSO')
    ax2.set_title("%s's Yearly DSO Performance in Comparison with Total DSO" %sheets[i])
    ax2.set_xticks([p+0.1 for p in pos])
    ax2.set_xticklabels(period)
    
    
    # Set the range of the y-axis for better usage
    plt.ylim(60 , 130)
    
    
    plt.grid()
    plt.legend(['Total', '%s' % sheets[i]])
    fig2.savefig("bar%s.png" %i, dpi=72)
    pdf.savefig(fig2)
    
    
    
    
    # Risk pie
    # Labels and values for the chart
    labels3 = [sheets[i], 'Rest of the A/R']
    x3 = customers[i].iloc[len(customers[i]) -1, 4]/ar_total
    sizes3 = [x3, 1-x3]
    
    explode = (0.1, 0) 
    fig3, ax3 = plt.subplots()
    ax3.pie(sizes3, explode=explode, colors=['tan', 'lightblue'], autopct='%1.1f%%', startangle=15)
    ax3.axis('equal')
    plt.figure(figsize=(4,2))
    
    
    # Title and legend
    ax3.set_title("%s's A/R ratio" %sheets[i])
    ax3.legend(labels3, loc='best', bbox_to_anchor=(0.3, 0.2,))
    fig3.savefig("risk_pie%s.png" % i, dpi = 72)
    pdf.savefig(fig3)
    
    
    
    
    # LoG pie
    # Labels and values for the chart
    labels4 = [sheets[i], 'Rest of the credit limit']
    x4 = customers[i].iloc[len(customers[i]) -1, 2]/credit_lim_total
    sizes4 = [x4, 1-x4]
    
    explode = (0.1, 0) 
    fig4, ax4 = plt.subplots()
    ax4.pie(sizes4, explode=explode, colors=['tan', 'lightblue'], autopct='%1.1f%%', startangle=15)
    ax4.axis('equal')
    plt.figure(figsize=(4,2))
    
    
    # Title and legend
    ax4.set_title("%s's Credit Limit Ratio" %sheets[i])
    ax4.legend(labels4, loc='best', bbox_to_anchor=(0.3, 0.2,))
    fig4.savefig("cl_pie%s.png" % i, dpi = 72)
    pdf.savefig(fig4)
    
    
    
    
    # Add to Excel file
    img1 = openpyxl.drawing.image.Image('sales_pie%s.png' % i)
    ws.add_image(img1, ws.cell(row=max_rows+4, column=1).coordinate)
    img2 = openpyxl.drawing.image.Image('bar%s.png' %i)
    ws.add_image(img2, ws.cell(row=16, column=16).coordinate)
    img3 = openpyxl.drawing.image.Image('risk_pie%s.png' %i)
    ws.add_image(img3, ws.cell(row=max_rows+4, column=7).coordinate)   # Change rows 
    img4 = openpyxl.drawing.image.Image('cl_pie%s.png' %i)
    ws.add_image(img4, ws.cell(row=max_rows+20, column=1).coordinate)   # Rows same as img3, change columns
    
    
wb.save('%s.xlsx' % report_path)
pdf.close()




# Measure the time elapsed from the start
end = time.time()
duration = end-start
print('Report is generated in %s seconds' %duration)
################################ END ##########################################
